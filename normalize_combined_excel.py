import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
import os

# Input and output files
INPUT_FILE = 'Combined_Excels.xlsx'
OUTPUT_FILE = 'Normalized_Combined_Excels.xlsx'

def check_file_exists(file_path):
    """
    Check if a file exists and provide helpful error message
    """
    if not os.path.exists(file_path):
        print(f"\nError: File '{file_path}' not found!")
        print("Please make sure you have run the copy_excels_to_sheets.py script first")
        print("to create the Combined_Excels.xlsx file.")
        return False
    return True

def get_all_headers(wb):
    """
    Get all unique headers across all sheets
    """
    all_headers = set()
    for sheet_name in wb.sheet_names:
        df = pd.read_excel(INPUT_FILE, sheet_name=sheet_name, nrows=1)
        all_headers.update(df.columns)
    return sorted(list(all_headers))

def normalize_combined_excel():
    print("Starting to normalize combined Excel file...")
    
    # Check if input file exists
    if not check_file_exists(INPUT_FILE):
        return
    
    try:
        # Get all unique headers across sheets
        wb = pd.ExcelFile(INPUT_FILE)
        all_headers = get_all_headers(wb)
        print(f"\nFound {len(all_headers)} unique columns across all sheets")
        
        # Create a new workbook
        wb_normalized = Workbook()
        # Remove the default sheet
        wb_normalized.remove(wb_normalized.active)
        
        # Process each sheet
        for sheet_name in wb.sheet_names:
            print(f"\nProcessing sheet: {sheet_name}")
            
            # Read the sheet into pandas
            df = pd.read_excel(INPUT_FILE, sheet_name=sheet_name)
            
            # Create a new DataFrame with all headers
            df_normalized = pd.DataFrame(columns=all_headers)
            
            # Copy data from original DataFrame
            for col in df.columns:
                if col in all_headers:
                    df_normalized[col] = df[col]
            
            # Create a new worksheet
            ws_normalized = wb_normalized.create_sheet(title=sheet_name)
            
            # Write headers
            for col_idx, header in enumerate(all_headers, 1):
                cell = ws_normalized.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
            
            # Write data
            for row_idx, row in df_normalized.iterrows():
                for col_idx, value in enumerate(row, 1):
                    cell = ws_normalized.cell(row=row_idx + 2, column=col_idx, value=value)
            
            print(f"Completed processing sheet: {sheet_name}")
        
        # Save the normalized workbook
        print(f"\nSaving normalized file as {OUTPUT_FILE}...")
        wb_normalized.save(OUTPUT_FILE)
        print("Normalization completed successfully!")
        
    except Exception as e:
        print(f"\nAn error occurred: {str(e)}")
        print("Please make sure the Combined_Excels.xlsx file is not open in Excel")
        print("and that you have write permissions in the current directory.")

if __name__ == '__main__':
    normalize_combined_excel() 