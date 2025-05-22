import openpyxl
import os

def merge_spanning_cells(file_path):
    """Process an Excel file and merge cells containing text that spans multiple cells."""
    print(f"Processing file: {file_path}")
    
    # Load the workbook
    wb = openpyxl.load_workbook(file_path)
    
    # Process each sheet
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"Processing sheet: {sheet_name}")
        
        # Get the dimensions of the sheet
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        # Dictionary to store text and their locations
        text_locations = {}
        
        # First pass: collect all text and their locations
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                # Skip if cell is part of a merged range
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    continue
                    
                if cell.value:
                    text = str(cell.value).strip()
                    if text:
                        if text not in text_locations:
                            text_locations[text] = []
                        text_locations[text].append((row, col))
        
        # Second pass: merge cells for text that appears in multiple cells
        for text, locations in text_locations.items():
            if len(locations) > 1:
                # Sort locations by row and column
                locations.sort()
                
                # Get the range of cells to merge
                start_row = min(loc[0] for loc in locations)
                end_row = max(loc[0] for loc in locations)
                start_col = min(loc[1] for loc in locations)
                end_col = max(loc[1] for loc in locations)
                
                # Create a copy of merged ranges to avoid modification during iteration
                merged_ranges = list(sheet.merged_cells.ranges)
                
                # Unmerge any existing merged cells in the range
                for merged_range in merged_ranges:
                    if (merged_range.min_row <= end_row and merged_range.max_row >= start_row and
                        merged_range.min_col <= end_col and merged_range.max_col >= start_col):
                        sheet.unmerge_cells(str(merged_range))
                
                # Merge the cells
                sheet.merge_cells(
                    start_row=start_row,
                    start_column=start_col,
                    end_row=end_row,
                    end_column=end_col
                )
                
                # Set the value in the top-left cell of the merged range
                sheet.cell(row=start_row, column=start_col).value = text
                
                print(f"Merged cells for text: {text}")
    
    # Create output directory if it doesn't exist
    output_dir = os.path.dirname(file_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Save the modified workbook
    output_path = os.path.join(output_dir, f"merged_{os.path.basename(file_path)}")
    wb.save(output_path)
    print(f"Saved merged file as: {output_path}")

if __name__ == "__main__":
    # Replace 'your_file.xlsx' with your actual Excel file name
    file_path = 'All-Excels/001-IR-E-U-0456.xlsx'
    merge_spanning_cells(file_path)
