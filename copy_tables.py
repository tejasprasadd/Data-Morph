import openpyxl
from openpyxl.utils import get_column_letter
import os
from copy import copy

def detect_table_boundaries(sheet, start_row, start_col):
    """Detect table boundaries by looking for continuous data blocks"""
    # Initialize boundaries
    end_row = start_row
    end_col = start_col
    
    # Find the last row of the table
    while end_row <= sheet.max_row:
        # Check if this row has any non-empty cells
        row_has_data = False
        for col in range(start_col, sheet.max_column + 1):
            cell_value = sheet.cell(row=end_row, column=col).value
            if cell_value is not None and str(cell_value).strip() != '':
                row_has_data = True
                break
        if not row_has_data:
            break
        end_row += 1
    end_row -= 1
    
    # Find the last column of the table
    while end_col <= sheet.max_column:
        # Check if this column has any non-empty cells
        col_has_data = False
        for row in range(start_row, end_row + 1):
            cell_value = sheet.cell(row=row, column=end_col).value
            if cell_value is not None and str(cell_value).strip() != '':
                col_has_data = True
                break
        if not col_has_data:
            break
        end_col += 1
    end_col -= 1
    
    return end_row, end_col

def is_valid_table(sheet, start_row, start_col, end_row, end_col):
    """Check if the detected range is a valid table"""
    # A valid table should have at least 2 rows and 2 columns
    if end_row - start_row < 1 or end_col - start_col < 1:
        return False
    
    # Check if there's enough data density
    total_cells = (end_row - start_row + 1) * (end_col - start_col + 1)
    non_empty_cells = 0
    
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value is not None and str(cell_value).strip() != '':
                non_empty_cells += 1
    
    # At least 30% of cells should have data
    return (non_empty_cells / total_cells) >= 0.3

def copy_tables_rowwise():
    # Input and output file paths
    input_file = os.path.join('All-Excels', '001-IR-E-U-0456.xlsx')
    output_file = 'tables_rowwise.xlsx'
    
    # Load the source workbook
    source_wb = openpyxl.load_workbook(input_file)
    source_sheet = source_wb.active
    
    # Create a new workbook for output
    target_wb = openpyxl.Workbook()
    target_sheet = target_wb.active
    
    # Initialize row counter for target sheet
    current_row = 1
    
    # Find and copy tables
    used_ranges = []
    row = 1
    while row <= source_sheet.max_row:
        col = 1
        while col <= source_sheet.max_column:
            cell_value = source_sheet.cell(row=row, column=col).value
            if cell_value is not None and str(cell_value).strip() != '':
                # Try to detect table boundaries
                end_row, end_col = detect_table_boundaries(source_sheet, row, col)
                
                # Check if this is a valid table
                if is_valid_table(source_sheet, row, col, end_row, end_col):
                    # Check if this range overlaps with any existing range
                    range_overlaps = False
                    for start_row, start_col, end_row_prev, end_col_prev in used_ranges:
                        if not (end_row < start_row or row > end_row_prev or
                               end_col < start_col or col > end_col_prev):
                            range_overlaps = True
                            break
                    
                    if not range_overlaps:
                        # Copy the table to the target sheet
                        for r in range(row, end_row + 1):
                            for c in range(col, end_col + 1):
                                source_cell = source_sheet.cell(row=r, column=c)
                                target_cell = target_sheet.cell(row=current_row + (r - row),
                                                              column=c - col + 1)
                                
                                # Copy value
                                target_cell.value = source_cell.value
                                
                                # Copy formatting properties individually
                                if source_cell.has_style:
                                    # Font
                                    if source_cell.font:
                                        target_cell.font = copy(source_cell.font)
                                    
                                    # Border
                                    if source_cell.border:
                                        target_cell.border = copy(source_cell.border)
                                    
                                    # Fill
                                    if source_cell.fill:
                                        target_cell.fill = copy(source_cell.fill)
                                    
                                    # Number format
                                    if source_cell.number_format:
                                        target_cell.number_format = source_cell.number_format
                                    
                                    # Alignment
                                    if source_cell.alignment:
                                        target_cell.alignment = copy(source_cell.alignment)
                        
                        # Add minimal spacing between tables (just one row)
                        current_row += (end_row - row + 1) + 1
                        
                        # Add this range to used ranges
                        used_ranges.append((row, col, end_row, end_col))
                        
                        # Skip to the end of this table
                        col = end_col
            col += 1
        row += 1
    
    # Save the new workbook
    target_wb.save(output_file)
    print(f"Tables have been copied to {output_file}")

if __name__ == "__main__":
    copy_tables_rowwise() 