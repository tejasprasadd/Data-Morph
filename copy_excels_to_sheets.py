import os
import re
from openpyxl import Workbook, load_workbook

# Directory containing the Excel files
EXCEL_DIR = 'All-Excels'
# Output file name
OUTPUT_FILE = 'Combined_Excels.xlsx'

def get_sorted_excel_files(directory):
    files = [f for f in os.listdir(directory) if f.endswith('.xlsx')]
    # Sort by the numeric prefix
    files.sort(key=lambda x: int(x.split('-')[0]))
    return files

def extract_sheet_name(filename):
    # Remove the numeric prefix and dash
    match = re.match(r'\d+-(.*)\.xlsx', filename)
    if match:
        return match.group(1)
    return filename.replace('.xlsx', '')

def copy_excel_sheets():
    print("Starting to process Excel files...")
    files = get_sorted_excel_files(EXCEL_DIR)
    print(f"Found {len(files)} Excel files to process")
    
    wb_out = Workbook()
    # Remove the default sheet
    default_sheet = wb_out.active
    wb_out.remove(default_sheet)

    for index, file in enumerate(files, 1):
        print(f"\nProcessing file {index}/{len(files)}: {file}")
        file_path = os.path.join(EXCEL_DIR, file)
        wb_in = load_workbook(file_path, data_only=False)  # Keep formulas as formulas
        sheet_in = wb_in.active  # Assume only one sheet per file
        sheet_name = extract_sheet_name(file)
        print(f"Creating sheet: {sheet_name}")
        ws_out = wb_out.create_sheet(title=sheet_name)
        
        # Simple copy-paste of all cells
        print("Copying all cells...")
        for row in sheet_in.iter_rows():
            for cell in row:
                ws_out[cell.coordinate].value = cell.value
        
        print(f"Completed processing {file}")

    print(f"\nSaving combined Excel file as {OUTPUT_FILE}...")
    wb_out.save(OUTPUT_FILE)
    print("Process completed successfully!")

if __name__ == '__main__':
    copy_excel_sheets() 