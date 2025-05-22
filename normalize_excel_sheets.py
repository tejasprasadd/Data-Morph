import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import numpy as np

# Directory containing the Excel files
EXCEL_DIR = 'All-Excels'
# Output directory for normalized files
OUTPUT_DIR = 'Normalized-Excels'

def get_sorted_excel_files(directory):
    files = [f for f in os.listdir(directory) if f.endswith('.xlsx')]
    # Sort by the numeric prefix
    files.sort(key=lambda x: int(x.split('-')[0]))
    return files

def extract_sheet_name(filename):
    # Remove the numeric prefix and dash
    match = re.match(r'\d+-(.*)\.xlsx', filename)
    if match:
        return match.group(1)
    return filename.replace('.xlsx', '')

def normalize_data(df):
    """
    Normalize the data in the DataFrame
    """
    # Create a copy to avoid modifying the original
    df_normalized = df.copy()
    
    # Convert all text to string and strip whitespace
    for col in df_normalized.columns:
        if df_normalized[col].dtype == 'object':
            df_normalized[col] = df_normalized[col].astype(str).str.strip()
    
    # Replace empty strings and 'nan' with actual NaN
    df_normalized = df_normalized.replace(['', 'nan', 'None', 'none', 'NULL', 'null'], np.nan)
    
    # Convert numeric columns to appropriate types
    for col in df_normalized.columns:
        try:
            # Try to convert to numeric, if possible
            df_normalized[col] = pd.to_numeric(df_normalized[col], errors='ignore')
        except:
            pass
    
    return df_normalized

def highlight_missing_values(worksheet):
    """
    Highlight missing values in the worksheet
    """
    # Define the fill for missing values
    missing_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    
    # Iterate through all cells
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is None or str(cell.value).strip() == '':
                cell.fill = missing_fill

def process_excel_file(file_path, output_path):
    """
    Process a single Excel file
    """
    print(f"Processing file: {os.path.basename(file_path)}")
    
    # Load the workbook
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Convert to DataFrame for easier processing
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)
    
    df = pd.DataFrame(data[1:], columns=data[0])
    
    # Normalize the data
    df_normalized = normalize_data(df)
    
    # Create a new workbook for the normalized data
    wb_normalized = load_workbook(file_path)
    ws_normalized = wb_normalized.active
    
    # Clear existing data
    for row in ws_normalized.iter_rows():
        for cell in row:
            cell.value = None
    
    # Write the normalized data back to the worksheet
    for r_idx, row in enumerate(df_normalized.itertuples(), 2):
        for c_idx, value in enumerate(row[1:], 1):
            ws_normalized.cell(row=r_idx, column=c_idx, value=value)
    
    # Write headers
    for c_idx, col_name in enumerate(df_normalized.columns, 1):
        ws_normalized.cell(row=1, column=c_idx, value=col_name)
        # Make headers bold
        ws_normalized.cell(row=1, column=c_idx).font = Font(bold=True)
    
    # Highlight missing values
    highlight_missing_values(ws_normalized)
    
    # Save the normalized workbook
    wb_normalized.save(output_path)
    print(f"Saved normalized file: {os.path.basename(output_path)}")

def normalize_excel_sheets():
    print("Starting to normalize Excel files...")
    
    # Create output directory if it doesn't exist
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
    
    files = get_sorted_excel_files(EXCEL_DIR)
    print(f"Found {len(files)} Excel files to process")
    
    for index, file in enumerate(files, 1):
        print(f"\nProcessing file {index}/{len(files)}: {file}")
        file_path = os.path.join(EXCEL_DIR, file)
        output_path = os.path.join(OUTPUT_DIR, file)
        process_excel_file(file_path, output_path)
    
    print("\nNormalization completed successfully!")

if __name__ == '__main__':
    normalize_excel_sheets() 